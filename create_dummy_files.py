import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import json


def create_dummy_pdf(content: str, filename: str):
    """Create a dummy PDF file with the given content."""
    # For now, we'll create a text file that represents a PDF
    # In a real implementation, you'd use a library like reportlab to create actual PDFs
    with open(filename, 'w') as f:
        f.write(f"DUMMY PDF CONTENT: {content}\n\n")
        f.write("This is a placeholder file for testing purposes.\n")
        f.write("In a real scenario, this would be an actual PDF document.\n")


def create_dummy_budget_xlsx(filename: str):
    """Create a dummy budget Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    
    # Add headers
    headers = ["Item", "Description", "Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add budget items
    budget_items = [
        ["Direct Labor", "Principal Investigator", 50000],
        ["Direct Labor", "Research Assistant", 30000],
        ["Materials", "Equipment and supplies", 20000],
        ["Travel", "Conference attendance", 5000],
        ["Subcontracts", "External consultants", 15000],
        ["Indirect Costs", "Overhead (50%)", 60000],
        ["TABA", "Technical and Business Assistance", 10000]
    ]
    
    for row, (item, desc, amount) in enumerate(budget_items, 2):
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=amount)
    
    # Add totals
    ws.cell(row=len(budget_items) + 2, column=1, value="Total Direct Costs")
    ws.cell(row=len(budget_items) + 2, column=3, value=120000)
    
    ws.cell(row=len(budget_items) + 3, column=1, value="Total Indirect Costs")
    ws.cell(row=len(budget_items) + 3, column=3, value=60000)
    
    ws.cell(row=len(budget_items) + 4, column=1, value="TOTAL")
    ws.cell(row=len(budget_items) + 4, column=3, value=180000)
    ws.cell(row=len(budget_items) + 4, column=3).font = Font(bold=True)
    
    # Set specific cells for the parser
    ws["C12"] = 15000  # Subcontract total
    ws["C17"] = 180000  # Total
    ws["C21"] = 10000   # TABA
    
    wb.save(filename)


def create_dummy_files(output_dir: Path):
    """Create all required dummy files."""
    output_dir.mkdir(exist_ok=True)
    
    print(f"Creating dummy files in {output_dir}")
    
    # Technical proposal
    tech_content = """
    TECHNICAL PROPOSAL
    
    Project Title: Advanced AI-Powered Proposal Evaluation System
    
    Executive Summary:
    This project aims to develop an innovative AI-powered system for automated proposal evaluation. 
    The system will leverage natural language processing and machine learning to assess technical 
    merit, commercial potential, and compliance of SBIR/STTR proposals.
    
    Technical Approach:
    Our approach combines state-of-the-art transformer models with domain-specific knowledge 
    extraction techniques. The system will analyze proposal text for technical feasibility, 
    innovation level, and implementation methodology.
    
    Innovation:
    The key innovation lies in our multi-modal evaluation framework that considers both 
    technical content and commercial viability. We employ advanced NLP techniques including 
    BERT-based classification and semantic similarity analysis.
    
    Risk Assessment:
    Primary risks include data quality issues and model interpretability. We mitigate these 
    through robust data preprocessing and explainable AI techniques.
    
    Implementation Plan:
    Phase I will focus on prototype development and validation. Phase II will involve 
    full-scale deployment and commercialization.
    
    Team Qualifications:
    Our team includes experts in AI/ML, proposal evaluation, and SBIR/STTR processes. 
    The PI has 15+ years of experience in machine learning and natural language processing.
    """
    
    create_dummy_pdf(tech_content, output_dir / "tech_proposal.pdf")
    print("Created tech_proposal.pdf")
    
    # Commercial proposal
    commercial_content = """
    COMMERCIAL PROPOSAL
    
    Market Analysis:
    The proposal evaluation market is valued at $2.5 billion annually, with significant 
    growth potential in government contracting. Current manual evaluation processes are 
    time-consuming and expensive.
    
    Customer Base:
    Primary customers include government agencies, research institutions, and private 
    companies that regularly evaluate proposals. The SBIR/STTR program alone processes 
    thousands of proposals annually.
    
    Value Proposition:
    Our system reduces evaluation time by 80% while improving consistency and accuracy. 
    This translates to significant cost savings and better decision-making.
    
    Competitive Analysis:
    Current solutions are limited to basic text analysis. Our AI-powered approach provides 
    deeper insights and more accurate assessments than existing tools.
    
    Commercialization Strategy:
    We will initially target SBIR/STTR programs, then expand to other government and 
    private sector markets. Revenue model includes licensing fees and subscription services.
    
    Transition Plan:
    Phase III will focus on commercialization through partnerships with government 
    contractors and direct sales to agencies.
    """
    
    create_dummy_pdf(commercial_content, output_dir / "commercial_proposal.pdf")
    print("Created commercial_proposal.pdf")
    
    # Team bios
    team_content = """
    TEAM BIOS
    
    Principal Investigator - Dr. Jane Smith
    Dr. Smith holds a Ph.D. in Computer Science from MIT and has 15+ years of experience 
    in artificial intelligence and machine learning. She has published 50+ peer-reviewed 
    papers and holds 10 patents in NLP applications.
    
    Co-PI - Dr. John Doe
    Dr. Doe is an expert in proposal evaluation with 20+ years of experience in SBIR/STTR 
    programs. He has served as a reviewer for multiple federal agencies and understands 
    the evaluation criteria and processes.
    
    Technical Lead - Sarah Johnson
    Ms. Johnson has 8 years of experience in software development and AI system design. 
    She has led multiple successful AI projects and specializes in natural language 
    processing applications.
    
    Business Development - Mike Wilson
    Mr. Wilson has 12 years of experience in business development and commercialization. 
    He has successfully launched 5 products in the government contracting space.
    """
    
    create_dummy_pdf(team_content, output_dir / "team_bios.pdf")
    print("Created team_bios.pdf")
    
    # Past performance
    past_performance_content = """
    PAST PERFORMANCE
    
    Company Background:
    Our company has successfully completed 15+ SBIR/STTR projects over the past 10 years, 
    with a 100% Phase I to Phase II transition rate. We have delivered innovative solutions 
    in AI, cybersecurity, and data analytics.
    
    Relevant Projects:
    1. AI-Powered Document Analysis System (Phase II SBIR, 2022-2024)
       - Developed automated document classification system
       - Achieved 95% accuracy in document categorization
       - Successfully commercialized to 3 government agencies
    
    2. Machine Learning for Risk Assessment (Phase I SBIR, 2021)
       - Developed prototype risk assessment tool
       - Demonstrated 85% accuracy in identifying high-risk proposals
       - Received Phase II award
    
    3. Natural Language Processing for Compliance (Phase II STTR, 2020-2022)
       - Built compliance checking system for government contracts
       - Processed 10,000+ documents with 90% accuracy
       - Licensed to 5 major contractors
    
    Key Personnel Experience:
    - PI: 15+ years in AI/ML, 10+ SBIR/STTR projects
    - Co-PI: 20+ years in proposal evaluation, former federal reviewer
    - Technical team: 25+ combined years in software development
    """
    
    create_dummy_pdf(past_performance_content, output_dir / "past_performance.pdf")
    print("Created past_performance.pdf")
    
    # Budget file
    create_dummy_budget_xlsx(output_dir / "budget.xlsx")
    print("Created budget.xlsx")
    
    print("\nAll dummy files created successfully!")
    print("You can now run the grading system with these files.")


def main():
    parser = argparse.ArgumentParser(description="Create dummy files for proposal testing")
    parser.add_argument("--output-dir", default="documents/proposal", type=Path, 
                       help="Output directory for dummy files")
    parser.add_argument("--overwrite", action="store_true", 
                       help="Overwrite existing files")
    
    args = parser.parse_args()
    
    if not args.overwrite:
        # Check if files already exist
        existing_files = []
        for file in ["tech_proposal.pdf", "commercial_proposal.pdf", "team_bios.pdf", 
                    "past_performance.pdf", "budget.xlsx"]:
            if (args.output_dir / file).exists():
                existing_files.append(file)
        
        if existing_files:
            print(f"Warning: The following files already exist: {', '.join(existing_files)}")
            print("Use --overwrite to replace them.")
            return
    
    create_dummy_files(args.output_dir)


if __name__ == "__main__":
    main() 